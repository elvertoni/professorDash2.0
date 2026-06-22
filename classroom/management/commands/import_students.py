import os
import openpyxl
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.conf import settings
from catalog.models import Disciplina
from classroom.models import Turma, Matricula
from accounts.models import User, AlunoProfile

# Maps spreadsheet discipline labels to catalog slugs in database
DISCIPLINA_MAP = {
    'Analise E Metodo Para Sistemas': 'analise-e-metodos-para-sistemas',
    'Analise E Projeto De Sistemas': 'analise-e-projeto-de-sistemas',
    'Inovacao Tec E Empreend': 'inovacao-tecnologia-e-empreendedorismo',
    'Introd A Computacao': 'introducao-a-computacao',
    'Programacao Front-End': 'programacao-front-end',
    'Programacao No Des De Sistemas': 'programacao-no-desenvolvimento-de-sistemas',
}

class Command(BaseCommand):
    help = 'Importa turmas e matricula estudantes a partir de uma planilha do Google Sheets (formato Excel).'

    def add_arguments(self, parser):
        parser.add_argument('file_path', type=str, help='Caminho do arquivo Excel (.xlsx)')
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Executa todo o processo em uma transação e desfaz as alterações no final.'
        )

    def handle(self, *args, **options):
        file_path = options['file_path']
        dry_run = options['dry_run']

        if not os.path.exists(file_path):
            raise CommandError(f"Arquivo não encontrado: {file_path}")

        self.stdout.write(f"Abrindo planilha: {file_path}")
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True)
        except Exception as e:
            raise CommandError(f"Erro ao ler arquivo Excel: {e}")

        if 'Resumo' not in wb.sheetnames:
            raise CommandError("A planilha deve conter uma aba chamada 'Resumo'.")

        resumo_sheet = wb['Resumo']

        # Find or fallback professor
        try:
            professor = User.objects.get(email='elvertoni@gmail.com')
        except User.DoesNotExist:
            professor = User.objects.filter(role=User.Role.ADMIN).first() or User.objects.filter(is_superuser=True).first()
            if not professor:
                raise CommandError("Nenhum usuário professor ('elvertoni@gmail.com') ou Admin encontrado no banco.")

        self.stdout.write(f"Professor responsável pelas turmas: {professor.nome_completo} ({professor.email})")

        created_users = 0
        updated_users = 0
        created_enrollments = 0
        updated_enrollments = 0
        skipped_rows = 0

        # Helper to match sheet name by truncated prefix
        def get_sheet_by_name(wb, full_name):
            for name in wb.sheetnames:
                if name == full_name:
                    return wb[name]
                # Excel sheets are max 31 chars, test first 25 chars
                if len(name) <= 31 and full_name.startswith(name[:25]):
                    return wb[name]
            return None

        # Execute inside transaction
        with transaction.atomic():
            if dry_run:
                self.stdout.write(self.style.WARNING("MODO DRY-RUN ATIVO - Nenhuma alteração será salva no banco."))

            # Iterate Resumo to find class and disciplines
            row_count = 0
            for row in resumo_sheet.iter_rows(min_row=2, values_only=True):
                # Columns: Turma, Disciplina, Alunos, Aba
                if not row or not any(cell is not None for cell in row[:4]):
                    continue

                turma_nome = str(row[0]).strip() if row[0] is not None else None
                disciplina_nome = str(row[1]).strip() if row[1] is not None else None
                aba_nome = str(row[3]).strip() if row[3] is not None else None

                if not turma_nome or not disciplina_nome or not aba_nome:
                    continue

                row_count += 1
                self.stdout.write(f"\nProcessando turma: {turma_nome} | Disciplina: {disciplina_nome}")

                # Map discipline
                disc_slug = DISCIPLINA_MAP.get(disciplina_nome)
                if not disc_slug:
                    self.stdout.write(self.style.ERROR(f"Aviso: Disciplina '{disciplina_nome}' não mapeada. Pulando."))
                    continue

                try:
                    disciplina = Disciplina.objects.get(slug=disc_slug)
                except Disciplina.DoesNotExist:
                    self.stdout.write(self.style.ERROR(f"Erro: Disciplina com slug '{disc_slug}' não cadastrada. Pulando."))
                    continue

                # Determine grade/serie
                if '3º' in turma_nome or '3' in turma_nome:
                    serie = '3º ANO'
                elif '2º' in turma_nome or '2' in turma_nome:
                    serie = '2º ANO'
                else:
                    serie = '1º ANO'

                # Find existing Turma by discipline and academic year 2026, or create
                turma, created = Turma.objects.get_or_create(
                    disciplina=disciplina,
                    ano_letivo=2026,
                    defaults={
                        'nome': turma_nome,
                        'serie': serie,
                        'professor': professor,
                        'ativa': True,
                    }
                )

                if not created:
                    # Update name to spreadsheet name and ensure it is active
                    if turma.nome != turma_nome or not turma.ativa:
                        turma.nome = turma_nome
                        turma.ativa = True
                        turma.save(update_fields=['nome', 'ativa', 'updated_at'])
                        self.stdout.write(f"Turma existente ID {turma.id} atualizada para o nome '{turma_nome}'.")
                else:
                    self.stdout.write(f"Nova Turma ID {turma.id} criada: '{turma_nome}'.")

                # Find the sheet in the workbook
                student_sheet = get_sheet_by_name(wb, aba_nome)
                if student_sheet is None:
                    self.stdout.write(self.style.ERROR(f"Aba '{aba_nome}' não encontrada no arquivo Excel. Pulando."))
                    continue

                # Process students
                sheet_student_count = 0
                for s_row in student_sheet.iter_rows(min_row=2, values_only=True):
                    # Columns: Nome, Email
                    if not s_row or not any(cell is not None for cell in s_row[:2]):
                        continue

                    s_nome = str(s_row[0]).strip() if s_row[0] is not None else None
                    s_email = str(s_row[1]).strip().lower() if s_row[1] is not None else None

                    if not s_nome or not s_email:
                        skipped_rows += 1
                        continue

                    # Create or update student User
                    user, u_created = User.objects.get_or_create(
                        email=s_email,
                        defaults={
                            'nome_completo': s_nome,
                            'role': User.Role.ALUNO,
                        }
                    )

                    if u_created:
                        user.set_unusable_password()
                        user.save()
                        created_users += 1
                    else:
                        # Ensure fields match
                        changed = False
                        if user.nome_completo != s_nome:
                            user.nome_completo = s_nome
                            changed = True
                        if user.role != User.Role.ALUNO:
                            user.role = User.Role.ALUNO
                            changed = True
                        if changed:
                            user.save(update_fields=['nome_completo', 'role', 'updated_at'])
                            updated_users += 1

                    # Aluno Profile
                    profile, p_created = AlunoProfile.objects.get_or_create(user=user)
                    if profile.grade != turma.serie:
                        profile.grade = turma.serie
                        profile.save(update_fields=['grade', 'updated_at'])

                    # Matricula
                    matricula, m_created = Matricula.objects.update_or_create(
                        turma=turma,
                        aluno=user,
                        defaults={
                            'status': Matricula.Status.ATIVA,
                        }
                    )

                    if m_created:
                        created_enrollments += 1
                    else:
                        if matricula.status != Matricula.Status.ATIVA:
                            matricula.status = Matricula.Status.ATIVA
                            matricula.save(update_fields=['status', 'updated_at'])
                        updated_enrollments += 1

                    sheet_student_count += 1

                self.stdout.write(f"Alunos processados na aba: {sheet_student_count}")

            if dry_run:
                transaction.set_rollback(True)
                self.stdout.write(self.style.WARNING("\n[DRY-RUN] Alterações revertidas com sucesso."))
            else:
                self.stdout.write(self.style.SUCCESS("\n[SUCESSO] Importação gravada com sucesso no banco de dados."))

        # Print report
        self.stdout.write("\n=== RELATÓRIO DE IMPORTAÇÃO ===")
        self.stdout.write(f"Novos usuários (alunos) criados: {created_users}")
        self.stdout.write(f"Usuários existentes atualizados: {updated_users}")
        self.stdout.write(f"Novas matrículas realizadas: {created_enrollments}")
        self.stdout.write(f"Matrículas existentes confirmadas ativas: {updated_enrollments}")
        self.stdout.write(f"Linhas ignoradas: {skipped_rows}")
        self.stdout.write("================================")
